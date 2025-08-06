import os
import pandas as pd
from datetime import datetime
from pynput import mouse, keyboard
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from mappings import combine_mapping

# ----------准备阶段----------
records = []                  # 操作记录
exit_flag = False             # 退出标志
start_time = datetime.now()   # 开始时间

# ----------鼠标点击事件处理----------
def on_click(x, y, button, pressed):
    event_type = '按下' if pressed else '松开'
    button = "左键" if str(button) == "Button.left" else "右键"
    print(f"鼠标{event_type}：{x}, {y}, {button}")

    if event_type == '按下':
        time_diff = (datetime.now() - start_time).total_seconds()
        record = {'时间': round(time_diff, 2), '事件': f'鼠标点击', '细节': f'{x},{y},{button}'}
        records.append(record)

# ----------键盘按键事件处理----------
def on_press(key):
    try:
        print(f"按键按下：'{key.char}'")
        time_diff = (datetime.now() - start_time).total_seconds()
        record = {'时间': round(time_diff, 2), '事件': '按键按下', '细节': repr(key.char)}
        records.append(record)

    except AttributeError:
        print(f"特殊按键按下：{key}")
        time_diff = (datetime.now() - start_time).total_seconds()
        record = {'时间': round(time_diff, 2), '事件': '特殊按键按下', '细节': repr(key)}
        records.append(record)

def on_release(key):
    global exit_flag
    print(f"按键松开：{key}")
    time_diff = (datetime.now() - start_time).total_seconds()
    record = {'时间': round(time_diff, 2), '事件': '按键松开', '细节': repr(key)}
    records.append(record)
    
    if key == keyboard.Key.esc:
        print("Esc 按下，正在结束监听器...")
        exit_flag = True  # 设置退出标志
        return False      # 返回 False 会停止监听器
    return True           # 默认返回 True，继续监听

# ----------启动鼠标和键盘监听器的函数----------
def start_listeners():
    with mouse.Listener(on_click=on_click) as mouse_listener,\
         keyboard.Listener(on_press=on_press, on_release=on_release) as keyboard_listener:
        
        while not exit_flag:   # 监听直到 exit_flag 为 True
            pass
        
        # 停止监听器
        mouse_listener.stop()
        keyboard_listener.stop()

# ----------主程序----------
start_listeners()     # 启动监听器

# ----------将记录保存到 Excel----------
for i in range(len(records)):
    if "按键" in records[i]["事件"]:
        detail = records[i]["细节"]
        detail = detail.replace('\\', '/')
        detail = detail[1 : -1]
        if ":" in detail:
            detail = detail.split(":")[0]
        elif detail in combine_mapping:
            detail = combine_mapping[detail]
        records[i]["细节"] = detail
    records[i]["条件"] = ""

idx = 1
excel_name = f'操作1.xlsx'
while os.path.exists(excel_name):
    idx += 1
    excel_name = f'操作{idx}.xlsx'

df = pd.DataFrame(records)
df.to_excel(excel_name, index=False)

# ----------设置 Excel 格式----------
wb = load_workbook(excel_name)
ws = wb.active

# 设置居中对齐
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 设置列宽
column_widths = {'A': 25, 'B': 25, 'C': 35, 'D': 75}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 保存格式化后的 Excel 文件
wb.save(excel_name)
print(f"操作记录已保存到 {excel_name}")
